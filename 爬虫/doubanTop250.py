#!/usr/bin/python3
# -*- coding : utf-8 -*-
# @Time : 2020/10/12 11:06
# @Author : Cheng
# @File : doubanTop250.py
# @Software : PyCharm

import requests
import bs4
import openpyxl

def getHTMLText(url):
    try:
        r = requests.get(url,headers = {'User-agent':'Mozilla/5.0'})
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        print('获取网页异常')

def fillflist(html):
    try:
        flist = []
        soup = bs4.BeautifulSoup(html,'html.parser')
        for li in soup.find('ol',class_='grid_view').find_all('li'):
            rank = li.find('em').string
            score = li.find('span',class_='rating_num').string
            name = li.find('span',class_='title').string
            info = li.find('p',class_='').text
            flist.append([rank,name,score,info.replace('\n','').replace(' ','')])
        return flist
    except:
        print('填充异常')

# def printfilmlist(flist):
#     try:
#         for f in flist:
#             print('{0:^10}{1:{4}^15}{2:^10}{3:{4}^10}'.format(f[0],f[1],f[2],f[3],chr(12288)))
#     except:
#         print('打印错误')

def save_to_excel(flist,i):
    global wb,ws
    if i == 0:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '序号'
        ws['B1'] = '电影名称'
        ws['C1'] = '评分'
        ws['D1'] = '简介'
    for f in flist:
        ws.append(f)
    if i == 9:
        wb.save('C:\\Users\\陌离\\Desktop\\豆瓣Top250.xlsx')

def main():
    for i in range(10):
        url = 'https://movie.douban.com/top250?start=' + str(25*i)
        html = getHTMLText(url)
        flist = fillflist(html)
        # printfilmlist(flist)
        save_to_excel(flist,i)

if __name__ == '__main__':
    main()