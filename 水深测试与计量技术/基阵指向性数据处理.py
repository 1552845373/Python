#!/usr/bin/python3
# -*- coding : utf-8 -*-
# @Time : 2021/5/14 17:31
# @Author : Cheng
# @File : 基阵指向性数据处理.py
# @Software : PyCharm


import xlrd
import openpyxl
import heapq
from numpy import *

delta_1_list, delta_2_list = [], []
for i in range(36):
    wb = xlrd.open_workbook("E:\\哈尔滨工程大学\\研一课程\\水声测试与计量技术\\实验数据\\水听器基阵指向性数据（水深30多）\\"+str(i*10)+".xls")
    table = wb.sheets()[0]
    CH1X = [table.row_values(i+1)[0] for i in range(table.nrows-1)]
    max_number_1 = heapq.nlargest(10, CH1X)
    max_1 = mean(max_number_1)
    CH1Y = [table.row_values(i+1)[1] for i in range(table.nrows-1)]
    min_number_1 = heapq.nsmallest(10, CH1Y)
    min_1 = mean(min_number_1)
    CH2X = [table.row_values(i+1)[2] for i in range(table.nrows-1)]
    max_number_2 = heapq.nlargest(10, CH1X)
    max_2 = mean(max_number_2)
    CH1X = [table.row_values(i+1)[3] for i in range(table.nrows-1)]
    min_number_2 = heapq.nsmallest(10, CH1X)
    min_2 = mean(min_number_2)
    delta_1 = max_1 - min_1
    delta_2 = max_2 - min_2
    delta_1_list.append(delta_1)
    delta_2_list.append(delta_2)

wb = openpyxl.load_workbook("E:\\哈尔滨工程大学\\研一课程\\水声测试与计量技术\\实验数据\\基阵指向性.xlsx")
ws = wb["Sheet1"]
for i in range(36):
    ws['B' + str(i + 1)] = delta_1_list[i]
    ws['C' + str(i + 1)] = delta_2_list[i]
wb.save("E:\\哈尔滨工程大学\\研一课程\\水声测试与计量技术\\实验数据\\基阵指向性.xlsx")
