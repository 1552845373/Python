#!/usr/bin/python3
# -*- coding : utf-8 -*-
# @Time : 2021/3/23 20:02
# @Author : Cheng
# @File : 统计学生提交作业.py
# @Software : PyCharm

import os
import re
import openpyxl

def file_name(file_dir):
    """
    提取文件夹中所有文件对应的学号
    :param file_dir: 文件夹路径
    :return: 学号列表
    """
    school_num = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            match = re.search(r'[A-Z]+\d*', file, re.I) # 使用正则表达式匹配文件名中以字母开头，后面接一串数字的部分。（re.I：忽视大小写）
            school_num.append(match.group(0))
    return school_num

def match(file_path, school_num):
    """
    匹配学生名单，并记录作业提交情况，保存匹配修改后的excel文件
    :param file_path: 学生名单路径
    :param school_num: 学号列表
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]
    col0 = [item.value for item in list(ws.columns)[0]]
    for i in range(89):
        ws['D'+str(i+2)] = '?' # 先对整体遍历，作业提交一栏全部写入“？”
    for num in school_num:
        index = col0.index(num) + 1
        ws['D'+str(index)] = '√' # 对学号列表进行遍历，如果有则表示交了，对应表格中提交一栏用“√“替换”？“
    wb.save(file_path)

if __name__ == '__main__':
    school_num = file_name('E:\\哈尔滨工程大学\\研一课程\\JS海洋学导论\\2')
    match('E:\\哈尔滨工程大学\\研一课程\\JS海洋学导论\\JS海洋学作业提交.xlsx', school_num)
